"""
Regulatory Risk Report Generator
=================================
Generates comprehensive Excel reports with risk analysis results.
Includes formatted tables, charts, and regulatory interpretations.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, Any, Optional
import io


class RiskReportGenerator:
    """
    Generates regulatory-compliant risk analysis reports in Excel format.
    """
    
    # Regulatory regime definitions
    REGULATORY_REGIMES = {
        'basel_iii': {
            'name': 'Basel III Market Risk',
            'var_confidence': 0.99,
            'var_horizon': 10,
            'stress_testing': True
        },
        'frtb': {
            'name': 'FRTB (Fundamental Review of Trading Book)',
            'var_confidence': 0.975,  # ES instead of VaR
            'var_horizon': 10,
            'stress_testing': True
        },
        'ucits': {
            'name': 'UCITS Global Exposure',
            'var_confidence': 0.99,
            'var_horizon': 20,  # Monthly
            'leverage_limit': 2.0
        },
        'emir': {
            'name': 'EMIR Risk Management',
            'var_confidence': 0.99,
            'var_horizon': 1,
            'daily_calculation': True
        },
        'mifid_ii': {
            'name': 'MiFID II Product Governance',
            'var_confidence': 0.95,
            'target_market': True
        }
    }
    
    def __init__(self, risk_results: Dict[str, Any], regime: str = 'basel_iii'):
        """
        Initialize report generator.
        
        Args:
            risk_results: Dictionary containing all risk analysis results
            regime: Regulatory regime to use for interpretation
        """
        self.risk_results = risk_results
        self.regime = regime
        self.regime_config = self.REGULATORY_REGIMES.get(regime, self.REGULATORY_REGIMES['basel_iii'])
    
    def generate_excel_report(self, output_path: Optional[str] = None) -> str:
        """
        Generate comprehensive Excel risk report.
        
        Args:
            output_path: Path to save the report. If None, generates timestamped filename.
            
        Returns:
            Path to the generated report
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"regulatory_risk_report_{timestamp}.xlsx"
        
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # Create all sheets
        self._create_cover_sheet(writer)
        self._create_summary_sheet(writer)
        self._create_moments_sheet(writer)
        self._create_correlation_sheet(writer)
        self._create_var_sheet(writer)
        self._create_pca_sheet(writer)
        self._create_clustering_sheet(writer)
        self._create_regulatory_interpretation(writer)
        
        writer.close()
        
        # Apply formatting
        self._format_workbook(output_path)
        
        return output_path
    
    def _create_cover_sheet(self, writer: pd.ExcelWriter):
        """Create cover page with report metadata."""
        df = pd.DataFrame({
            'Field': [
                'Report Title',
                'Generated Date',
                'Regulatory Regime',
                'Analysis Period',
                'Number of Assets',
                'Confidence Level',
                'Time Horizon',
                'Report Version'
            ],
            'Value': [
                'Automated Regulatory Risk Analysis Report',
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                self.regime_config['name'],
                'Historical Data Analysis',
                len(self.risk_results.get('statistical_moments', {}).keys()) - 1,  # Exclude portfolio
                f"{self.regime_config['var_confidence']*100:.0f}%",
                f"{self.regime_config.get('var_horizon', 1)} day(s)",
                '1.0'
            ]
        })
        df.to_excel(writer, sheet_name='Cover', index=False)
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter):
        """Create executive summary sheet."""
        portfolio_moments = self.risk_results.get('statistical_moments', {}).get('portfolio', {})
        var_results = self.risk_results.get('var_analysis', {}).get('historical', {})
        pca_results = self.risk_results.get('pca_analysis', {})
        
        summary_data = {
            'Metric Category': [],
            'Metric': [],
            'Value': [],
            'Interpretation': []
        }
        
        # Statistical moments
        summary_data['Metric Category'].extend(['Statistical Moments'] * 4)
        summary_data['Metric'].extend(['Mean Return (Daily)', 'Volatility', 'Skewness', 'Excess Kurtosis'])
        summary_data['Value'].extend([
            f"{portfolio_moments.get('mean', 0):.4f}",
            f"{portfolio_moments.get('std_dev', 0):.4f}",
            f"{portfolio_moments.get('skewness', 0):.4f}",
            f"{portfolio_moments.get('excess_kurtosis', 0):.4f}"
        ])
        summary_data['Interpretation'].extend([
            'Daily average return',
            'Risk measure (standard deviation)',
            'Negative = left tail risk',
            '>0 = fat tails, extreme events more likely'
        ])
        
        # VaR metrics
        summary_data['Metric Category'].extend(['Risk Metrics'] * 3)
        summary_data['Metric'].extend(['VaR 95%', 'VaR 99%', 'CVaR 99%'])
        summary_data['Value'].extend([
            f"{var_results.get('var_95', 0):.4f}",
            f"{var_results.get('var_99', 0):.4f}",
            f"{var_results.get('cvar_99', 0):.4f}"
        ])
        summary_data['Interpretation'].extend([
            'Max loss at 95% confidence',
            'Max loss at 99% confidence',
            'Expected loss beyond VaR 99%'
        ])
        
        # PCA metrics
        explained_var = pca_results.get('explained_variance_ratio', [0])
        cumvar = pca_results.get('cumulative_variance_ratio', [0])
        
        summary_data['Metric Category'].extend(['Factor Analysis'] * 3)
        summary_data['Metric'].extend(['PC1 Explained Var', 'PC1+PC2 Explained', 'Number of PCs'])
        summary_data['Value'].extend([
            f"{explained_var[0]*100:.2f}%" if len(explained_var) > 0 else "N/A",
            f"{cumvar[1]*100:.2f}%" if len(cumvar) > 1 else "N/A",
            str(pca_results.get('n_components', 0))
        ])
        summary_data['Interpretation'].extend([
            'Main risk factor contribution',
            'Top 2 factors combined',
            'Principal components retained'
        ])
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Executive Summary', index=False)
    
    def _create_moments_sheet(self, writer: pd.ExcelWriter):
        """Create detailed statistical moments sheet."""
        moments = self.risk_results.get('statistical_moments', {})
        
        data = {
            'Asset': [],
            'Mean': [],
            'Std Dev': [],
            'Variance': [],
            'Skewness': [],
            'Kurtosis': [],
            'Excess Kurtosis': []
        }
        
        for asset, metrics in moments.items():
            data['Asset'].append(asset)
            data['Mean'].append(metrics.get('mean', 0))
            data['Std Dev'].append(metrics.get('std_dev', 0))
            data['Variance'].append(metrics.get('variance', 0))
            data['Skewness'].append(metrics.get('skewness', 0))
            data['Kurtosis'].append(metrics.get('kurtosis', 0))
            data['Excess Kurtosis'].append(metrics.get('excess_kurtosis', 0))
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Statistical Moments', index=False)
    
    def _create_correlation_sheet(self, writer: pd.ExcelWriter):
        """Create correlation matrix sheet."""
        corr_matrix = self.risk_results.get('correlation_matrix', {})
        if corr_matrix:
            df = pd.DataFrame(corr_matrix)
            df.to_excel(writer, sheet_name='Correlation Matrix')
    
    def _create_var_sheet(self, writer: pd.ExcelWriter):
        """Create VaR analysis sheet with multiple methods."""
        var_analysis = self.risk_results.get('var_analysis', {})
        
        data = {
            'Method': [],
            'VaR 95%': [],
            'VaR 99%': [],
            'VaR 99.9%': [],
            'CVaR 95%': [],
            'CVaR 99%': [],
            'Time Horizon': []
        }
        
        for method, results in var_analysis.items():
            if isinstance(results, dict) and 'error' not in results:
                data['Method'].append(method.replace('_', ' ').title())
                data['VaR 95%'].append(results.get('var_95', 0))
                data['VaR 99%'].append(results.get('var_99', 0))
                data['VaR 99.9%'].append(results.get('var_99_9', 0))
                data['CVaR 95%'].append(results.get('cvar_95', 0))
                data['CVaR 99%'].append(results.get('cvar_99', 0))
                data['Time Horizon'].append(f"{results.get('time_horizon', 1)} day(s)")
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='VaR Analysis', index=False)
        
        # Add VaR comparison chart data
        if len(data['Method']) > 0:
            chart_data = pd.DataFrame({
                'Confidence Level': ['95%', '99%', '99.9%'],
                'Historical': [
                    var_analysis.get('historical', {}).get('var_95', 0),
                    var_analysis.get('historical', {}).get('var_99', 0),
                    var_analysis.get('historical', {}).get('var_99_9', 0)
                ],
                'Parametric': [
                    var_analysis.get('parametric', {}).get('var_95', 0),
                    var_analysis.get('parametric', {}).get('var_99', 0),
                    var_analysis.get('parametric', {}).get('var_99_9', 0)
                ],
                'Cornish-Fisher': [
                    var_analysis.get('cornish_fisher', {}).get('var_95', 0),
                    var_analysis.get('cornish_fisher', {}).get('var_99', 0),
                    var_analysis.get('cornish_fisher', {}).get('var_99_9', 0)
                ]
            })
            chart_data.to_excel(writer, sheet_name='VaR Comparison', index=False)
    
    def _create_pca_sheet(self, writer: pd.ExcelWriter):
        """Create PCA analysis sheet."""
        pca_results = self.risk_results.get('pca_analysis', {})
        
        # Explained variance
        explained_var = pca_results.get('explained_variance_ratio', [])
        cumvar = pca_results.get('cumulative_variance_ratio', [])
        
        if len(explained_var) > 0:
            data = {
                'Principal Component': [f'PC{i+1}' for i in range(len(explained_var))],
                'Explained Variance Ratio': explained_var,
                'Cumulative Variance': cumvar
            }
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name='PCA Analysis', index=False)
            
            # Loadings
            loadings = pca_results.get('loadings', {})
            if loadings:
                loadings_df = pd.DataFrame(loadings)
                loadings_df.to_excel(writer, sheet_name='PCA Loadings')
    
    def _create_clustering_sheet(self, writer: pd.ExcelWriter):
        """Create cluster analysis sheet."""
        cluster_results = self.risk_results.get('cluster_analysis', {})
        
        clusters = cluster_results.get('clusters', {})
        
        data = {
            'Cluster': [],
            'Assets': [],
            'Asset Count': []
        }
        
        for cluster_id, assets in clusters.items():
            data['Cluster'].append(f'Cluster {cluster_id + 1}')
            data['Assets'].append(', '.join(assets))
            data['Asset Count'].append(len(assets))
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Cluster Analysis', index=False)
        
        # Assignment mapping
        assignments = cluster_results.get('cluster_assignments', {})
        if assignments:
            assignment_df = pd.DataFrame({
                'Asset': list(assignments.keys()),
                'Cluster': [f"Cluster {v + 1}" for v in assignments.values()]
            })
            assignment_df.to_excel(writer, sheet_name='Cluster Assignments', index=False)
    
    def _create_regulatory_interpretation(self, writer: pd.ExcelWriter):
        """Create regulatory interpretation sheet."""
        portfolio_moments = self.risk_results.get('statistical_moments', {}).get('portfolio', {})
        var_results = self.risk_results.get('var_analysis', {}).get('historical', {})
        
        interpretations = []
        
        # Skewness interpretation
        skew = portfolio_moments.get('skewness', 0)
        if skew < -0.5:
            interpretations.append({
                'Risk Factor': 'Return Distribution Skewness',
                'Finding': f'Negative skewness ({skew:.3f})',
                'Regulatory Implication': 'Left tail risk detected. Under Basel III/FRTB, this indicates potential for extreme losses beyond normal VaR estimates.',
                'Recommendation': 'Consider stress testing with larger shock scenarios. Monitor for tail risk hedging.'
            })
        elif skew > 0.5:
            interpretations.append({
                'Risk Factor': 'Return Distribution Skewness',
                'Finding': f'Positive skewness ({skew:.3f})',
                'Regulatory Implication': 'Right tail risk indicates potential for extreme gains.',
                'Recommendation': 'Standard risk models may be conservative. Validate against backtesting results.'
            })
        
        # Kurtosis interpretation
        kurt = portfolio_moments.get('excess_kurtosis', 0)
        if kurt > 1.0:
            interpretations.append({
                'Risk Factor': 'Tail Risk (Kurtosis)',
                'Finding': f'High excess kurtosis ({kurt:.3f})',
                'Regulatory Implication': 'Fat tails indicate higher probability of extreme events than predicted by normal distribution.',
                'Recommendation': 'Use non-parametric VaR methods (historical simulation) or Cornish-Fisher adjustment. Required under FRTB for non-modellable risk factors.'
            })
        
        # VaR interpretation
        var_99 = var_results.get('var_99', 0)
        if abs(var_99) > 0.05:  # 5% daily VaR
            interpretations.append({
                'Risk Factor': 'Value at Risk',
                'Finding': f'99% VaR: {var_99:.4f} ({var_99*100:.2f}%)',
                'Regulatory Implication': 'High VaR indicates significant market risk exposure.',
                'Recommendation': 'Ensure sufficient capital allocation under Basel III market risk rules. Consider diversification or hedging strategies.'
            })
        
        # PCA interpretation
        pca_results = self.risk_results.get('pca_analysis', {})
        explained_var = pca_results.get('explained_variance_ratio', [])
        if len(explained_var) > 0 and explained_var[0] > 0.5:
            interpretations.append({
                'Risk Factor': 'Factor Concentration',
                'Finding': f'First PC explains {explained_var[0]*100:.1f}% of variance',
                'Regulatory Implication': 'High concentration in single risk factor indicates lack of diversification.',
                'Recommendation': 'Under UCITS/EMIR, review diversification requirements. Consider factor hedging.'
            })
        
        df = pd.DataFrame(interpretations)
        if len(df) > 0:
            df.to_excel(writer, sheet_name='Regulatory Interpretation', index=False)
        
        # Add regime-specific guidance
        regime_guidance = {
            'Regime': [self.regime_config['name']],
            'Applicable Regulation': [self._get_regulation_reference()],
            'Key Requirements': [self._get_key_requirements()],
            'Compliance Notes': [self._get_compliance_notes()]
        }
        regime_df = pd.DataFrame(regime_guidance)
        regime_df.to_excel(writer, sheet_name='Regime Guidance', index=False)
    
    def _get_regulation_reference(self) -> str:
        """Get regulation reference for current regime."""
        references = {
            'basel_iii': 'Basel Committee on Banking Supervision - Minimum Capital Requirements for Market Risk',
            'frtb': 'BCBS 352 - Fundamental Review of the Trading Book',
            'ucits': 'Directive 2009/65/EC - UCITS V',
            'emir': 'Regulation (EU) No 648/2012 - EMIR',
            'mifid_ii': 'Directive 2014/65/EU - MiFID II'
        }
        return references.get(self.regime, 'General Risk Management Standards')
    
    def _get_key_requirements(self) -> str:
        """Get key requirements for current regime."""
        requirements = {
            'basel_iii': '99% VaR, 10-day holding period, Stressed VaR, Backtesting',
            'frtb': '97.5% Expected Shortfall, 10-day horizon, NMRF identification',
            'ucits': '99% VaR, 1-month horizon, 2x leverage limit',
            'emir': 'Daily VaR calculation, 99% confidence, Model validation',
            'mifid_ii': 'Product governance, Target market assessment, Risk warnings'
        }
        return requirements.get(self.regime, 'Standard risk management practices')
    
    def _get_compliance_notes(self) -> str:
        """Get compliance notes for current regime."""
        return 'This report provides indicative risk metrics. Formal regulatory submissions require additional documentation and model validation.'
    
    def _format_workbook(self, filepath: str):
        """Apply formatting to the Excel workbook."""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = load_workbook(filepath)
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        title_font = Font(bold=True, size=14)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        wb.save(filepath)


def generate_sample_report(risk_results: Dict, output_path: str = None) -> str:
    """
    Generate a sample risk report.
    
    Args:
        risk_results: Dictionary containing risk analysis results
        output_path: Output file path
        
    Returns:
        Path to generated report
    """
    generator = RiskReportGenerator(risk_results, regime='basel_iii')
    return generator.generate_excel_report(output_path)
